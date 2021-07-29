print("{}".format("开发者：兰锦航").center(60,"="))
print("854649049@qq.com".center(67,"="))
print(" ")
print("".center(67,"#"))
print("请在使用前仔细阅读说明书".center(55,"#"))
print("".center(67,"#"))
print(" ")
print("请稍等，正在初始化程序并检查文件")
print(" ")
import openpyxl as px
import time as t

#simlpy the grammer read a cell
def READ(a):
    return sheet[str(a)].value

#simlpy the grammer write a cell
def WRITE(a,b):
    sheet2[str(upper(a))].value=b

AB="ABCDEFGHIJKLMNOPQRSTUVWXYZ"

title_name=[]
message=[]
#len==24
    

#load class grade

try:
    wb=px.load_workbook("grade.xlsx")
    sheet=wb.worksheets[0]
except:
    print("抱歉，程序错误，请关闭程序重新检查，多次出现问题请联系开发者")
    t.sleep(10)
else:
    print("加载成功，可以开始")

    linenum=int(input("请输入需要复制的列数（小于等于27的数字）："))
    rownum=int(input("请输入学生人数："))
    nameline=str(input("请输入需要用哪一列的信息来命名文件（学生名字在哪一列）{大写字母}："))
    nameline.upper()
    print("请稍等，正在处理......")
    try:
        #add title
        for i in range(0,linenum,1):
            title_name.append(READ(AB[i]+"1"))


        #add message
        def add_message(b):
            for i in range(0,linenum,1):
                message.append(READ(AB[i]+str(b)))
        ##ok lets begin##

        for j in range(0,rownum,1):
            temp=j+2
            add_message(temp)
        #    print(message)

            tepwb=px.Workbook()
            tepwb.create_sheet(index=j,title=str(j+1))
            sheet2=tepwb.worksheets[0]
            for k in range(0,linenum,1):
                sheet2[AB[k]+"1"].value=title_name[k]
                sheet2[AB[k]+"2"].value=message[k]
            tepwb.save(str(sheet[nameline+str(temp)].value)+".xlsx")
            message=[]
    except:
        print("抱歉，程序错误，请关闭程序重新检查，多次出现问题请联系开发者")
        t.sleep(10)
    else:
        print("批量生成成功，感谢使用")
        t.sleep(3)
            




       
