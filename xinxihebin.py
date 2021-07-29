# *** coding:utf-8 ***
"""
作者：LJH
日期：2021 07 23
"""
import openpyxl as px
need_A_info=['C']
appenda=['Q']
grade=''
student_id=''
information=[]
name_number_a = 0
name_number_b = 0
id_number_a = 0
id_number_b = 0
la = 5
lb = 2
# load two excel
try:
    wba = px.load_workbook("A.xlsx")
    wbb = px.load_workbook("B.xlsx")
    wbc = px.load_workbook("C.xlsx")
except:
    print("打开错误")
    print("")
else:
    print("打开成功")
    print("")
    sheeta = wba.worksheets[0]
    sheetb = wbb.worksheets[0]
    sheetc = wbc.worksheets[0]

    def robota(a):
        a.upper()
        return sheeta[str(a)].value


    def robotb(a):
        a.upper()
        return sheetb[str(a)].value

    def robotc(a):
        a.upper()
        return sheetc[str(a)].value

for i in range(2,227):
    student_id = robotb('B'+str(i))
    for j in range(2,231):
        if robotc('A'+str(j))==student_id:
            print('FOUND')
            for item in need_A_info:
                information.append(robotc(item+str(j)))
    for k in range(len(information)):
        sheetb[appenda[k]+str(i)].value=information[k]
    print(information)
    information.clear()
wbb.save('B.xlsx')


